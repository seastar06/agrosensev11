"""
AgroSense v4 — Sentinel-2 NDVI
- STAC Catalog ile en yakın tarih bulma (hızlı)
- O günü dar aralıkla çek → hızlı execute
- Gerçek tarih bilgisi göster
- Tümünü seç düzeltmesi
- Excel/CSV sayısal
"""
import streamlit as st
import folium
from folium.plugins import Draw
from streamlit_folium import st_folium
import pandas as pd
import requests
import json, io, zipfile, tempfile, os, math, time
from datetime import datetime, timedelta
from shapely.geometry import shape
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="AgroSense NDVI", page_icon="🌿",
                   layout="wide", initial_sidebar_state="expanded")
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=Syne:wght@700;800&display=swap');
html,body,[class*="css"]{font-family:'Space Mono',monospace;}
.stApp{background:#0a0e0a;color:#e2f0e2;}
.block-container{padding-top:1rem;}
h1,h2,h3{font-family:'Syne',sans-serif;color:#4ade80;}
.mcard{background:#111811;border:1px solid #1e2e1e;border-radius:8px;padding:12px;text-align:center;margin:4px 0;}
.mval{font-size:22px;font-weight:700;color:#4ade80;}
.mlbl{font-size:10px;color:#8aab8a;text-transform:uppercase;letter-spacing:1px;}
.dwarn{background:#fff3cd;border:1px solid #f59e0b;border-radius:6px;padding:8px 12px;color:#856404;font-size:12px;margin:8px 0;}
.dinfo{background:#d1ecf1;border:1px solid #0c5460;border-radius:6px;padding:8px 12px;color:#0c5460;font-size:12px;margin:8px 0;}
[data-testid="stSidebar"]{background:#111811;border-right:1px solid #1e2e1e;}
[data-testid="stSidebar"] *{color:#e2f0e2 !important;}
.stButton>button{background:#4ade80;color:#0a0e0a;border:none;font-family:'Space Mono',monospace;font-weight:700;border-radius:6px;width:100%;}
.stButton>button:hover{background:#22c55e;}
</style>
""", unsafe_allow_html=True)

# ── Kimlik ────────────────────────────────────────────────────
SH_CLIENT_ID     = "sh-645dcab9-79b1-42a2-9d89-edee62b45fe1"
SH_CLIENT_SECRET = "S2xXnoyA9RkciUtMv4DVntkV2hFQZbMd"
SH_WMS           = f"https://sh.dataspace.copernicus.eu/ogc/wms/{SH_CLIENT_ID}"
CDSE_TOKEN_URL   = "https://identity.dataspace.copernicus.eu/auth/realms/CDSE/protocol/openid-connect/token"
CDSE_STAC        = "https://catalogue.dataspace.copernicus.eu/stac/search"
OPENEO_URL       = "https://openeo.dataspace.copernicus.eu"

# ── Session state ─────────────────────────────────────────────
for k, v in {
    "features": [], "selected_ids": [], "ndvi_results": {},
    "ndvi_dates": [], "active_date": None,
    "map_center": [39.0, 35.0], "map_zoom": 6,
    "date_warnings": {}, "last_draw_count": 0,
    "do_select_all": False,
}.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ── Token (OAuth2 — requests ile, cache'li) ───────────────────
@st.cache_data(ttl=3000, show_spinner=False)
def get_token():
    r = requests.post(CDSE_TOKEN_URL, data={
        "grant_type":    "client_credentials",
        "client_id":     SH_CLIENT_ID,
        "client_secret": SH_CLIENT_SECRET,
    }, timeout=15)
    r.raise_for_status()
    return r.json()["access_token"]

# ── STAC: ±15 gün içinde en yakın bulutsuz sahne ─────────────
def find_nearest_scene(bbox, target_date_str, days=15):
    """
    CDSE STAC Catalog API ile en yakın Sentinel-2 sahnesini bul.
    Hızlı: sadece metadata, veri indirme yok.
    bbox: [west, south, east, north]
    """
    token = get_token()
    d     = datetime.strptime(target_date_str, "%Y-%m-%d")
    start = (d - timedelta(days=days)).strftime("%Y-%m-%dT00:00:00Z")
    end   = (d + timedelta(days=days)).strftime("%Y-%m-%dT23:59:59Z")

    body = {
        "collections": ["SENTINEL-2"],
        "bbox":        bbox,
        "datetime":    f"{start}/{end}",
        "limit":       50,
        "filter":      {"op":"lte","args":[{"property":"eo:cloud_cover"},70]},
        "filter-lang": "cql2-json",
    }
    r = requests.post(CDSE_STAC, json=body,
                      headers={"Authorization": f"Bearer {token}",
                                "Content-Type": "application/json"},
                      timeout=20)
    if not r.ok:
        # fallback: filtre olmadan dene
        body2 = {k: v for k, v in body.items() if k not in ("filter","filter-lang")}
        r = requests.post(CDSE_STAC, json=body2,
                          headers={"Authorization": f"Bearer {token}",
                                    "Content-Type": "application/json"},
                          timeout=20)
    if not r.ok:
        raise RuntimeError(f"STAC {r.status_code}: {r.text[:200]}")

    items = r.json().get("features", [])
    if not items:
        return None, None

    # En yakın tarihi seç
    items.sort(key=lambda x: abs(
        datetime.strptime(x["properties"]["datetime"][:10], "%Y-%m-%d") - d))
    best = items[0]
    actual_date = best["properties"]["datetime"][:10]
    return actual_date, best

# ── OpenEO: tek gün, tüm parseller batch ─────────────────────
@st.cache_resource(show_spinner=False)
def get_openeo():
    import openeo
    conn = openeo.connect(OPENEO_URL)
    conn.authenticate_oidc_client_credentials(
        client_id=SH_CLIENT_ID, client_secret=SH_CLIENT_SECRET)
    return conn

def fetch_ndvi_for_date(features, actual_date_str):
    """
    Kesin tarihi bilinen bir gün için tüm parsellerin NDVI'sini çek.
    Tek gün → aggregate_spatial → çok hızlı.
    """
    conn = get_openeo()
    geoms = [shape(f["geom"]) for f in features]
    u = geoms[0]
    for g in geoms[1:]: u = u.union(g)
    b = u.bounds

    # Sadece o günü yükle (+1 gün buffer)
    d_end = (datetime.strptime(actual_date_str, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")

    cube = conn.load_collection(
        "SENTINEL2_L2A",
        spatial_extent={"west":b[0],"south":b[1],"east":b[2],"north":b[3]},
        temporal_extent=[actual_date_str, d_end],
        bands=["B04","B08"],
        max_cloud_cover=90,
    )
    b08  = cube.band("B08")
    b04  = cube.band("B04")
    ndvi = (b08 - b04) / (b08 + b04)

    # Zaman boyutunu kaldır (tek gün var zaten)
    ndvi = ndvi.reduce_dimension(dimension="t", reducer="mean")

    fc = {
        "type": "FeatureCollection",
        "features": [
            {"type":"Feature","id":f["id"],"geometry":f["geom"],"properties":{"fid":f["id"]}}
            for f in features
        ]
    }

    result = ndvi.aggregate_spatial(geometries=fc, reducer="mean")
    raw    = result.execute()

    # Debug: raw yanıtı kaydet
    st.session_state["_last_raw"] = str(raw)[:500]

    return parse_openeo_response(raw, features)

def parse_openeo_response(raw, features):
    """
    OpenEO aggregate_spatial çıktısı:
    - Liste: [[v1], [v2], ...]  veya [v1, v2, ...]
    - Dict:  {"2024-06-01T..": [[v1],[v2],...]}  (zaman serisi)
    Her parsel için tek sayı çıkar.
    """
    out = {f["id"]: None for f in features}

    def first_num(x):
        if isinstance(x, (int, float)) and x == x and not (isinstance(x,float) and abs(x)>10):
            return float(x)
        if isinstance(x, list):
            for i in x:
                v = first_num(i)
                if v is not None: return v
        return None

    vals = []
    if isinstance(raw, list):
        # [[v],[v],...] veya [v,v,...]
        for item in raw:
            vals.append(first_num(item))
    elif isinstance(raw, dict):
        # Zaman serisi dict: keys = timestamps
        # Her key altında [[v1,v2,...]] (parsel başına) olabilir
        # Ya da flat list
        for ts_val in raw.values():
            if isinstance(ts_val, list):
                for item in ts_val:
                    vals.append(first_num(item))
            else:
                vals.append(first_num(ts_val))
            break  # Tek zaman dilimi var
    else:
        v = first_num(raw)
        if v is not None:
            vals = [v] * len(features)

    for i, feat in enumerate(features):
        v = vals[i] if i < len(vals) else None
        if v is not None and -1 <= v <= 1:
            out[feat["id"]] = round(v, 3)
        # else: None kalır

    return out

# ── Ana NDVI fonksiyonu: STAC + OpenEO ───────────────────────
def fetch_ndvi_batch(features, target_date_str):
    """
    1. STAC ile ±15 gün içinde en yakın tarihi bul (hızlı)
    2. O gün için tüm parselleri tek batch ile çek
    Returns: {fid: ndvi_val}, actual_date_str
    """
    # Tüm parsellerin bbox'u
    geoms = [shape(f["geom"]) for f in features]
    u = geoms[0]
    for g in geoms[1:]: u = u.union(g)
    b = u.bounds
    bbox = [b[0], b[1], b[2], b[3]]

    actual_date, scene = find_nearest_scene(bbox, target_date_str, days=15)
    if actual_date is None:
        return {f["id"]: None for f in features}, None

    ndvi_vals = fetch_ndvi_for_date(features, actual_date)
    return ndvi_vals, actual_date

# ── Yardımcılar ───────────────────────────────────────────────
def ndvi_color(v):
    if v is None: return "#888"
    if v<0:      return "#5ab4d6"
    if v<0.10:   return "#d73027"
    if v<0.20:   return "#f46d43"
    if v<0.25:   return "#fdae61"
    if v<0.35:   return "#fee08b"
    if v<0.45:   return "#d9ef8b"
    if v<0.55:   return "#a6d96a"
    if v<0.65:   return "#66bd63"
    return "#1a9850"

def ndvi_status(v):
    if v is None: return "Veri yok"
    if v>0.35:   return "🌾 Ekili"
    if v>0.15:   return "🟡 Geçiş"
    return "🌱 Boş/Nadas"

def area_dk(g):
    try:
        s = shape(g)
        if s.geom_type == "MultiPolygon": s = max(s.geoms, key=lambda x: x.area)
        c = list(s.exterior.coords)
        lat0 = sum(x[1] for x in c) / len(c)
        R=6371000; lm=R*math.pi/180; lo=R*math.cos(math.radians(lat0))*math.pi/180
        a=0; n=len(c)
        for i in range(n):
            x1,y1=c[i][0]*lo,c[i][1]*lm; x2,y2=c[(i+1)%n][0]*lo,c[(i+1)%n][1]*lm
            a+=x1*y2-x2*y1
        return round(abs(a)/2/1000,2)
    except: return 0

# ── Dosya okuma ───────────────────────────────────────────────
def parse_geojson(d):
    fc = d if d.get("type")=="FeatureCollection" else {"type":"FeatureCollection","features":[d]}
    return [{"id":str(i),"props":f.get("properties") or {},"geom":f["geometry"]}
            for i,f in enumerate(fc.get("features",[]))
            if f.get("geometry") and "Polygon" in f["geometry"].get("type","")]

def parse_kml(text):
    NS="http://www.opengis.net/kml/2.2"; root=ET.fromstring(text); out=[]; idx=0
    for pm in root.iter(f"{{{NS}}}Placemark"):
        ne=pm.find(f"{{{NS}}}name")
        props={"name":(ne.text or "") if ne is not None else ""}
        for dd in pm.iter(f"{{{NS}}}Data"):
            k=dd.get("name",""); ve=dd.find(f"{{{NS}}}value")
            if k and ve is not None: props[k]=ve.text or ""
        for poly in pm.iter(f"{{{NS}}}Polygon"):
            oc=poly.find(f".//{{{NS}}}outerBoundaryIs//{{{NS}}}coordinates")
            if oc is None or not oc.text: continue
            coords=[]
            for tok in oc.text.strip().split():
                p=tok.split(",")
                if len(p)>=2:
                    try: coords.append([float(p[0]),float(p[1])])
                    except: pass
            if len(coords)>=3:
                if coords[0]!=coords[-1]: coords.append(coords[0])
                out.append({"id":str(idx),"props":props,
                            "geom":{"type":"Polygon","coordinates":[coords]}}); idx+=1
    return out

def load_file(uf):
    ext=uf.name.rsplit(".",1)[-1].lower()
    if ext in("geojson","json"): return parse_geojson(json.loads(uf.read()))
    if ext=="kml": return parse_kml(uf.read().decode("utf-8","ignore"))
    if ext=="kmz":
        with zipfile.ZipFile(io.BytesIO(uf.read())) as z:
            kn=next((n for n in z.namelist() if n.lower().endswith(".kml")),None)
            if not kn: raise ValueError("KMZ içinde KML yok")
            return parse_kml(z.read(kn).decode("utf-8","ignore"))
    if ext=="zip":
        raw=uf.read()
        with tempfile.TemporaryDirectory() as td:
            with zipfile.ZipFile(io.BytesIO(raw)) as z: z.extractall(td)
            for fn in os.listdir(td):
                fp=os.path.join(td,fn)
                if fn.endswith(".shp"):
                    import shapefile as sf
                    r=sf.Reader(fp); fields=[f[0] for f in r.fields[1:]]
                    return [{"id":str(i),"props":dict(zip(fields,sr.record)),
                             "geom":sr.shape.__geo_interface__}
                            for i,sr in enumerate(r.shapeRecords())
                            if "Polygon" in sr.shape.__geo_interface__.get("type","")]
                if fn.endswith(".kml"):
                    return parse_kml(open(fp,encoding="utf-8",errors="ignore").read())
                if fn.endswith((".geojson",".json")):
                    return parse_geojson(json.load(open(fp,encoding="utf-8")))
    raise ValueError(f"Desteklenmeyen: {ext}")

# ── Harita ────────────────────────────────────────────────────
def build_map(feats, sel_ids, act_date, ndvi_res, sent_date):
    m = folium.Map(location=st.session_state.map_center,
                   zoom_start=st.session_state.map_zoom,
                   tiles="https://mt1.google.com/vt/lyrs=s&x={x}&y={y}&z={z}",
                   attr="Google Satellite", max_zoom=21)
    Draw(export=False,
         draw_options={"polyline":False,"circle":False,"marker":False,
                       "circlemarker":False,"rectangle":True,"polygon":True},
         edit_options={"edit":False,"remove":True}).add_to(m)
    if sent_date:
        folium.WmsTileLayer(url=SH_WMS, name="🛰 Sentinel RGB",
            layers="TRUE-COLOR", fmt="image/png", transparent=True, version="1.3.0",
            extra_params={"time":f"{sent_date}/{sent_date}","maxcc":80},
            opacity=0.85, overlay=True, control=True).add_to(m)
        folium.WmsTileLayer(url=SH_WMS, name="🌿 NDVI Overlay",
            layers="NDVI", fmt="image/png", transparent=True, version="1.3.0",
            extra_params={"time":f"{sent_date}/{sent_date}","maxcc":80},
            opacity=0.55, overlay=True, control=True).add_to(m)
    for feat in feats:
        fid=feat["id"]; is_sel=fid in sel_ids
        val=ndvi_res.get(fid,{}).get(act_date,{}).get("ndvi") if act_date else None
        fc=ndvi_color(val) if val is not None else ("#4ade80" if is_sel else "#22c55e")
        fo=0.65 if val is not None else (0.25 if is_sel else 0.08)
        tip=[f"<b>#{fid}</b>"]+[f"{k}: {v}" for k,v in list(feat["props"].items())[:5]]
        if val is not None:
            act=ndvi_res.get(fid,{}).get(act_date,{}).get("actual_date",act_date)
            tip+=[f"<b>NDVI: {val:.3f}</b>",f"<b>{ndvi_status(val)}</b>"]
            if act and act!=act_date: tip.append(f"📅 Gerçek: {act}")
        folium.GeoJson(feat["geom"],
            style_function=lambda x,fc=fc,fo=fo,is_sel=is_sel:{
                "fillColor":fc,"fillOpacity":fo,
                "color":"#ffff00" if is_sel else "#4ade80","weight":3 if is_sel else 1.5},
            tooltip=folium.Tooltip("<br>".join(tip))).add_to(m)
    folium.LayerControl().add_to(m)
    return m

# ── Export ────────────────────────────────────────────────────
def build_rows(feats, sel_ids, ndvi_res, dates):
    fm={f["id"]:f for f in feats}; rows=[]
    for fid in sel_ids:
        feat=fm.get(fid)
        if not feat: continue
        rec={"Parsel_#":fid}
        for k,v in feat["props"].items(): rec[str(k)]=v
        rec["Alan_Dekar"]=area_dk(feat["geom"])
        for date in sorted(dates):
            nd=ndvi_res.get(fid,{}).get(date,{}); val=nd.get("ndvi"); act=nd.get("actual_date",date)
            suffix=f"(ger:{act})" if act and act!=date else ""
            rec[f"NDVI_{date}{suffix}"]=val
            rec[f"Durum_{date}"]=("Ekili" if val is not None and val>0.35 else
                                   "Geçiş" if val is not None and val>0.15 else
                                   "Boş/Nadas" if val is not None else "")
        rows.append(rec)
    return rows

def ts_rows(feats, sel_ids, ndvi_res, dates):
    fm={f["id"]:f for f in feats}; rows=[]
    for fid in sel_ids:
        feat=fm.get(fid)
        if not feat: continue
        base={"Parsel_#":fid}
        for k,v in feat["props"].items(): base[str(k)]=v
        base["Alan_Dekar"]=area_dk(feat["geom"])
        for date in sorted(dates):
            nd=ndvi_res.get(fid,{}).get(date,{}); val=nd.get("ndvi"); act=nd.get("actual_date",date)
            row=dict(base); row["Hedef_Tarih"]=date; row["Gercek_Tarih"]=act or date
            row["NDVI"]=val
            row["Durum"]=("Ekili" if val is not None and val>0.35 else
                          "Geçiş" if val is not None and val>0.15 else
                          "Boş/Nadas" if val is not None else "")
            rows.append(row)
    return rows

def to_xlsx(rows, sheet="Analiz", hcol="1E5631"):
    if not rows: return io.BytesIO()
    df=pd.DataFrame(rows); buf=io.BytesIO()
    with pd.ExcelWriter(buf,engine="openpyxl") as w:
        df.to_excel(w,index=False,sheet_name=sheet)
        ws=w.sheets[sheet]; hf=PatternFill("solid",fgColor=hcol)
        for c in ws[1]:
            c.fill=hf; c.font=Font(color="FFFFFF",bold=True)
            c.alignment=Alignment(horizontal="center")
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width=min(
                max(len(str(c.value or "")) for c in col)+4,40)
    buf.seek(0); return buf

def to_csv(rows):
    if not rows: return io.BytesIO()
    buf=io.BytesIO(); buf.write("\ufeff".encode("utf-8"))
    buf.write(pd.DataFrame(rows).to_csv(index=False,float_format="%.4f").encode("utf-8"))
    buf.seek(0); return buf

# ══════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## 🌿 AgroSense")
    st.markdown("**Sentinel-2 NDVI Parsel Analiz**")
    st.markdown("---")

    # DOSYA
    st.markdown("### 📁 Dosya Yükle")
    uf=st.file_uploader("SHP(zip) · KML · KMZ · GeoJSON",
                        type=["zip","kml","kmz","geojson","json"],
                        label_visibility="collapsed")
    if uf:
        try:
            loaded=load_file(uf)
            st.session_state.features=loaded
            st.session_state.selected_ids=[]
            st.session_state.ndvi_results={}
            lons=[]; lats=[]
            for f in loaded:
                coords=f["geom"].get("coordinates",[])
                ring=coords[0] if coords else []
                for c in ring:
                    if isinstance(c,(list,tuple)) and len(c)>=2:
                        lons.append(c[0]); lats.append(c[1])
            if lons:
                st.session_state.map_center=[(min(lats)+max(lats))/2,(min(lons)+max(lons))/2]
                st.session_state.map_zoom=13
            st.success(f"✓ {len(loaded)} parsel yüklendi")
        except Exception as e: st.error(f"Hata: {e}")

    st.markdown("---")
    feats=st.session_state.features

    if feats:
        st.markdown("### ☑️ Parsel Seç")
        c1,c2=st.columns(2)
        with c1:
            if st.button("✅ Tümünü Seç"):
                st.session_state.selected_ids=[f["id"] for f in feats]
                st.rerun()
        with c2:
            if st.button("🗑 Temizle"):
                st.session_state.selected_ids=[]
                st.rerun()

        all_keys=list({k for f in feats for k in f["props"]})
        str_keys=[k for k in all_keys
                  if all(isinstance(f["props"].get(k),(str,type(None))) for f in feats)]
        if str_keys:
            fcol=st.selectbox("Filtre sütunu",["—"]+str_keys,key="fcol")
            if fcol!="—":
                uvals=sorted({str(f["props"].get(fcol,"")) for f in feats if f["props"].get(fcol)})
                svals=st.multiselect("Değer seç",uvals,key="fvals")
                if svals and st.button("Filtreyi Uygula"):
                    st.session_state.selected_ids=[f["id"] for f in feats
                                                   if str(f["props"].get(fcol,"")) in svals]
                    st.rerun()

        # Manuel multiselect — default her zaman session_state'ten
        lk=all_keys[0] if all_keys else None
        id_opts={f["id"]:f"#{f['id']} {str(f['props'].get(lk,''))[:18] if lk else ''}" for f in feats}
        sel=st.multiselect("Manuel seç",list(id_opts),
                           format_func=lambda x:id_opts[x],
                           default=st.session_state.selected_ids,
                           key="msel")
        # Sadece kullanıcı elle değiştirince güncelle
        if set(sel) != set(st.session_state.selected_ids):
            st.session_state.selected_ids=sel

        st.caption(f"Seçili: **{len(st.session_state.selected_ids)}** / {len(feats)}")

    st.markdown("---")

    # TARİHLER
    st.markdown("### 📅 NDVI Tarihleri")
    t1,t2=st.tabs(["Tek/Toplu","Aralık"])

    with t1:
        nd=st.date_input("Tarih",value=datetime.today(),key="di",label_visibility="collapsed")
        if st.button("➕ Ekle"):
            ds=nd.strftime("%Y-%m-%d")
            if ds not in st.session_state.ndvi_dates:
                st.session_state.ndvi_dates.append(ds)
            else: st.warning("Zaten var")
        bulk=st.text_input("Toplu (virgülle): 2024-06-01, 2024-07-15",
                           key="bulk",label_visibility="collapsed",
                           placeholder="2024-06-01, 2024-07-15, 2024-08-20")
        if st.button("➕ Toplu Ekle") and bulk:
            added=0
            for ds in bulk.split(","):
                ds=ds.strip()
                try:
                    datetime.strptime(ds,"%Y-%m-%d")
                    if ds not in st.session_state.ndvi_dates:
                        st.session_state.ndvi_dates.append(ds); added+=1
                except: pass
            if added: st.success(f"✓ {added} tarih eklendi")

    with t2:
        ca,cb=st.columns(2)
        with ca: rs=st.date_input("Başlangıç",value=datetime.today()-timedelta(days=90),key="rs",label_visibility="collapsed")
        with cb: re=st.date_input("Bitiş",value=datetime.today(),key="re",label_visibility="collapsed")
        iv=st.selectbox("Aralık (gün)",[5,10,15,30],index=2,key="iv")
        if st.button("➕ Aralık Ekle"):
            cur=rs; added=0
            while cur<=re:
                ds=cur.strftime("%Y-%m-%d")
                if ds not in st.session_state.ndvi_dates:
                    st.session_state.ndvi_dates.append(ds); added+=1
                cur+=timedelta(days=iv)
            st.success(f"✓ {added} tarih eklendi")

    if st.session_state.ndvi_dates:
        for dstr in list(st.session_state.ndvi_dates):
            r1,r2=st.columns([4,1])
            with r1:
                w=st.session_state.date_warnings.get(dstr)
                st.markdown(f"📅 {dstr}"+(f" → `{w}`" if w and w!=dstr else ""))
            with r2:
                if st.button("✕",key=f"rm_{dstr}"):
                    st.session_state.ndvi_dates.remove(dstr); st.rerun()
        if st.button("🗑 Tümünü Temizle"):
            st.session_state.ndvi_dates=[]; st.rerun()
        st.session_state.active_date=st.selectbox(
            "Aktif tarih (harita)",st.session_state.ndvi_dates,key="adsel")

    st.markdown("---")

    # ANALİZ
    st.markdown("### 🔬 NDVI Analiz")
    if st.button("◉ Analiz Başlat",type="primary"):
        sel=st.session_state.selected_ids
        dates=st.session_state.ndvi_dates
        if not sel: st.error("Önce parsel seçin")
        elif not dates: st.error("Önce tarih ekleyin")
        else:
            fm={f["id"]:f for f in st.session_state.features}
            sel_feats=[fm[fid] for fid in sel if fid in fm]
            # Hesaplanmamış tarih-parsel çiftleri
            to_do=[]
            for date in dates:
                missing=[f for f in sel_feats
                         if date not in st.session_state.ndvi_results.get(f["id"],{})]
                if missing: to_do.append((date,missing))

            if not to_do:
                st.info("Tüm değerler zaten hesaplanmış.")
            else:
                prog=st.progress(0,text="⏳ STAC → en yakın tarih bulunuyor...")
                errors=[]
                for i,(date,missing) in enumerate(to_do):
                    prog.progress(i/len(to_do),
                                  text=f"📡 {date} — {len(missing)} parsel batch...")
                    try:
                        vals, actual = fetch_ndvi_batch(missing, date)
                        if actual is None:
                            for f in missing:
                                if f["id"] not in st.session_state.ndvi_results:
                                    st.session_state.ndvi_results[f["id"]]={}
                                st.session_state.ndvi_results[f["id"]][date]={
                                    "ndvi":None,"actual_date":date}
                            errors.append(f"{date}: ±15 gün içinde görüntü yok")
                            continue
                        if actual != date:
                            st.session_state.date_warnings[date]=actual
                        for fid,val in vals.items():
                            if fid not in st.session_state.ndvi_results:
                                st.session_state.ndvi_results[fid]={}
                            st.session_state.ndvi_results[fid][date]={
                                "ndvi":val,"actual_date":actual}
                    except Exception as e:
                        errors.append(f"{date}: {str(e)[:120]}")
                        for f in missing:
                            if f["id"] not in st.session_state.ndvi_results:
                                st.session_state.ndvi_results[f["id"]]={}
                            st.session_state.ndvi_results[f["id"]][date]={
                                "ndvi":None,"actual_date":date}

                prog.progress(1.0,text="✓ Tamamlandı!")
                time.sleep(0.3); prog.empty()

                if errors:
                    for e in errors: st.warning(e)
                else:
                    st.success(f"✓ {len(sel)} parsel × {len(dates)} tarih!")

                # Debug: son raw yanıt
                if "_last_raw" in st.session_state:
                    with st.expander("🔍 Debug: son API yanıtı"):
                        st.code(st.session_state["_last_raw"])

                st.rerun()

    st.markdown("---")

    # EXPORT
    st.markdown("### 💾 Dışa Aktar")
    if st.session_state.selected_ids and st.session_state.ndvi_dates:
        etype=st.radio("Tür",["Parsel bazlı","Zaman serisi"],horizontal=True,key="etype")
        fmt=st.radio("Format",["xlsx","csv","ikisi de"],horizontal=True,key="efmt")
        if st.button("⬇️ İndir"):
            fa=st.session_state.features; sa=st.session_state.selected_ids
            na=st.session_state.ndvi_results; da=st.session_state.ndvi_dates
            fn=f"agrosense_{datetime.today().strftime('%Y%m%d_%H%M')}"
            if etype=="Parsel bazlı":
                rows=build_rows(fa,sa,na,da)
                if fmt in("xlsx","ikisi de"):
                    st.download_button("📊 Excel (Parsel)",data=to_xlsx(rows),
                        file_name=f"{fn}_parsel.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                if fmt in("csv","ikisi de"):
                    st.download_button("📄 CSV (Parsel)",data=to_csv(rows),
                        file_name=f"{fn}_parsel.csv",mime="text/csv")
            else:
                rows=ts_rows(fa,sa,na,da)
                if fmt in("xlsx","ikisi de"):
                    st.download_button("📊 Excel (Zaman Serisi)",
                        data=to_xlsx(rows,"Zaman_Serisi","1A3A6A"),
                        file_name=f"{fn}_zaman_serisi.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                if fmt in("csv","ikisi de"):
                    st.download_button("📄 CSV (Zaman Serisi)",data=to_csv(rows),
                        file_name=f"{fn}_zaman_serisi.csv",mime="text/csv")
    else: st.caption("Parsel seçin ve tarih ekleyin")

# ══════════════════════════════════════════════════════════════
# ANA ALAN
# ══════════════════════════════════════════════════════════════
st.markdown("# 🌿 AgroSense — Sentinel-2 NDVI")

feats=st.session_state.features
sel_ids=st.session_state.selected_ids
ndvi_res=st.session_state.ndvi_results
act_date=st.session_state.active_date

# Metrikler
if sel_ids and act_date:
    vals=[ndvi_res.get(fid,{}).get(act_date,{}).get("ndvi") for fid in sel_ids
          if ndvi_res.get(fid,{}).get(act_date,{}).get("ndvi") is not None]
    avg=round(sum(vals)/len(vals),3) if vals else None
    c1,c2,c3,c4=st.columns(4)
    c1.markdown(f'<div class="mcard"><div class="mval">{len(sel_ids)}</div><div class="mlbl">Seçili</div></div>',unsafe_allow_html=True)
    c2.markdown(f'<div class="mcard"><div class="mval" style="color:{ndvi_color(avg)}">{avg or "—"}</div><div class="mlbl">Ort. NDVI</div></div>',unsafe_allow_html=True)
    c3.markdown(f'<div class="mcard"><div class="mval" style="color:#1a9850">{sum(1 for v in vals if v>0.35)}</div><div class="mlbl">🌾 Ekili</div></div>',unsafe_allow_html=True)
    c4.markdown(f'<div class="mcard"><div class="mval" style="color:#d73027">{sum(1 for v in vals if v<=0.15)}</div><div class="mlbl">🌱 Boş/Nadas</div></div>',unsafe_allow_html=True)
    warn=st.session_state.date_warnings.get(act_date)
    if warn and warn!=act_date:
        st.markdown(f'<div class="dwarn">📅 <b>{act_date}</b> için görüntü yok — en yakın: <b>{warn}</b> kullanıldı (±15 gün)</div>',unsafe_allow_html=True)

# Harita
sent_date=act_date or (st.session_state.ndvi_dates[-1] if st.session_state.ndvi_dates else None)
m=build_map(feats,sel_ids,act_date,ndvi_res,sent_date)
map_out=st_folium(m,width="100%",height=550,returned_objects=["all_drawings"])

# Alan çizerek seçim
drawings=map_out.get("all_drawings") or []
if len(drawings)!=st.session_state.last_draw_count and drawings and feats:
    st.session_state.last_draw_count=len(drawings)
    dg=drawings[-1].get("geometry")
    if dg:
        try:
            ds=shape(dg)
            matched=[f["id"] for f in feats
                     if ds.contains(shape(f["geom"]).centroid) or ds.intersects(shape(f["geom"]))]
            if matched:
                ex=set(st.session_state.selected_ids); ex.update(matched)
                st.session_state.selected_ids=list(ex)
                st.success(f"✓ **{len(matched)}** parsel seçildi (toplam: {len(st.session_state.selected_ids)})")
                st.rerun()
            else: st.info("Çizilen alanda parsel yok")
        except Exception as e: st.warning(f"Seçim hatası: {e}")

# Zaman serisi grafik
if sel_ids and len(st.session_state.ndvi_dates)>1:
    has=any(ndvi_res.get(fid,{}).get(d,{}).get("ndvi") is not None
            for fid in sel_ids for d in st.session_state.ndvi_dates)
    if has:
        st.markdown("---")
        st.markdown("### 📈 Zaman Serisi")
        fm={f["id"]:f for f in feats}; ts={}
        for fid in sel_ids:
            feat=fm.get(fid); lbl=f"#{fid}"
            if feat and feat["props"]:
                k0=list(feat["props"].keys())[0]
                lbl=f"#{fid} {str(feat['props'][k0])[:15]}"
            s={d:ndvi_res.get(fid,{}).get(d,{}).get("ndvi")
               for d in sorted(st.session_state.ndvi_dates)
               if ndvi_res.get(fid,{}).get(d,{}).get("ndvi") is not None}
            if s: ts[lbl]=s
        if ts:
            df_ts=pd.DataFrame(ts).T
            df_ts.columns=pd.to_datetime(df_ts.columns)
            st.line_chart(df_ts.T)

# Tablo
if sel_ids and st.session_state.ndvi_dates:
    st.markdown("---")
    st.markdown("### 📋 Sonuçlar")
    rows=build_rows(feats,sel_ids,ndvi_res,st.session_state.ndvi_dates)
    if rows: st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)

with st.expander("🎨 NDVI Renk Skalası"):
    items=[("#5ab4d6","<0","Su"),("#d73027","0-0.10","Çıplak"),("#f46d43","0.10-0.20","Düşük"),
           ("#fdae61","0.20-0.25","Düşük+"),("#fee08b","0.25-0.35","Orta"),("#d9ef8b","0.35-0.45","İyi"),
           ("#a6d96a","0.45-0.55","İyi+"),("#66bd63","0.55-0.65","Yüksek"),("#1a9850",">0.65","Çok Yüksek")]
    cols=st.columns(9)
    for col,(color,rng,lbl) in zip(cols,items):
        tc="#fff" if color not in("#fee08b","#d9ef8b","#a6d96a") else "#222"
        col.markdown(f'<div style="background:{color};border-radius:6px;padding:8px;text-align:center;color:{tc}"><div style="font-size:10px;font-weight:700">{lbl}</div><div style="font-size:9px">{rng}</div></div>',unsafe_allow_html=True)
